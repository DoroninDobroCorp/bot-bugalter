import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


async def export_operations_to_excel(operations, start_date, end_date):
    # Создаем новую рабочую книгу Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "История операций"

    # Определяем заголовки столбцов
    headers = ["ID", "Дата", "Пользователь", "Тип операции", "Описание"]

    # Записываем заголовки в первую строку листа
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Записываем данные операций в остальные строки
    for row_num, operation in enumerate(operations, 2):
        sheet.cell(row=row_num, column=1, value=operation.id)
        sheet.cell(row=row_num, column=2,
                   value=operation.date.strftime('%d.%m.%Y %H:%M:%S'))
        sheet.cell(row=row_num, column=3, value=operation.user_name)
        sheet.cell(row=row_num, column=4, value=operation.operation_type)
        sheet.cell(row=row_num, column=5,
                   value=operation.operation_description)

    # Устанавливаем ширину столбцов по содержимому
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].auto_size = True

    # Формируем название файла с использованием start_date и end_date
    filename = f"history_{start_date.strftime('%d.%m.%Y')}_{end_date.strftime('%d.%m.%Y')}.xlsx"

    # Сохраняем рабочую книгу в буфер в памяти
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    return excel_file, filename


async def export_commissions_to_excel(commissions, start_date, end_date):
    # Создаем новую рабочую книгу Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "История комиссий"

    # Определяем заголовки столбцов
    headers = ["ID", "Дата", "Пользователь", "Сумма комиссии", "Тип комиссии",
               "Описание"]

    # Записываем заголовки в первую строку листа
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Записываем данные комиссий в остальные строки
    for row_num, commission in enumerate(commissions, 2):
        sheet.cell(row=row_num, column=1, value=commission.id)
        sheet.cell(row=row_num, column=2,
                   value=commission.date.strftime('%d.%m.%Y %H:%M:%S'))
        sheet.cell(row=row_num, column=3, value=commission.user_name)
        sheet.cell(row=row_num, column=4, value=commission.commission)
        sheet.cell(row=row_num, column=5, value=commission.commission_type)
        sheet.cell(row=row_num, column=6,
                   value=commission.commission_description)

    # Устанавливаем ширину столбцов по содержимому
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].auto_size = True

    # Формируем название файла с использованием start_date и end_date
    filename = f"commissions_{start_date.strftime('%d.%m.%Y')}_{end_date.strftime('%d.%m.%Y')}.xlsx"

    # Сохраняем рабочую книгу в буфер в памяти
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    return excel_file, filename
